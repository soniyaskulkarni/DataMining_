import openpyxl
import numpy as np
import math
# opening the excel sheet having data
wb_obj = openpyxl.load_workbook("K:/Final Year WCE/DataMiningLab/Normalization/bookip_1.xlsx")
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
n_col = sheet_obj.max_column
new_max = 200
new_min = 100
# traversing the excel sheet columnwise and processing each column
for j in range(2, n_col+1):
    tmp = []  # array to store values of column
    sum = 0
    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=j)
        tmp.append(cell_obj.value)
        sum += cell_obj.value
        temp = int(cell_obj.value)
    # mean of column
    mean_tmp = sum/(m_row-1)
    mx = max(tmp)
    mn = min(tmp)
    l = len(tmp)
    # calculating standard deviation of column
    std_d = 0
    a = 0
    while a < l:
        std_d = std_d + (mean_tmp-tmp[a])*(mean_tmp-tmp[a])
        a = a+1
    std_d = math.sqrt(std_d)
    std_d = std_d/math.sqrt(l)
    print(std_d)

    s = 0  # variable to traverse
    z_score = []  # array to mstore z_score
    while s < l:
        # normalization min_max
        tmp[s] = ((tmp[s]-mn)/(mx-mn))*(new_max-new_min) + new_min
        # normalization z_score
        z_score.append((tmp[s]-mean_tmp)/std_d)
        s = s+1
    c = 0
    t1 = sheet_obj.cell(row=1, column=j)
    title = sheet_obj.cell(row=1, column=j+n_col+1)
    title.value = t1.value+"_min_max"
    title = sheet_obj.cell(row=1, column=j+(2*n_col+2))
    title.value = t1.value+"_z_score"
    # writing values back to sheet
    for p in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row=p, column=j+n_col+1)
        cell_obj.value = tmp[c]
        cell_obj1 = sheet_obj.cell(row=p, column=j+(2*n_col+2))
        cell_obj1.value = z_score[c]
        c = c+1
    wb_obj.save("K:/Final Year WCE/DataMiningLab/Normalization/bookop_1.xlsx")